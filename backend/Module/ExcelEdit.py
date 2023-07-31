from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv


class ExcelEdit:
    def __init__(self):
        self.wb = None
        self.worksheet = None
        self.RowLists = []
        self.work_folder = ""

    def WriteAll(self):
        for idx, elem in enumerate(self.RowLists):
            self.WriteRow(idx + 1, elem)

    def WriteRow(self, row_num, text_list):
        for idx, elem in enumerate(text_list):
            self.worksheet.cell(row=row_num, column=idx + 1).value = elem

    def SaveExcel(self, result_dir):
        self.wb.save(f"./{result_dir}/items.xlsm")

    def AppendRowList(self, val):
        self.RowLists.append(val)

    def SetWorkFolder(self, val):
        self.work_folder = val
        self.wb = load_workbook(
            f'./{self.work_folder}/items.xlsm',
            keep_vba=True)
        self.worksheet = self.wb['items']

    def CSV2Sheet(self):
        csv_filename = f'./{self.work_folder}/items.csv'
        excel_filename = f'./{self.work_folder}/items.xlsm'
        with open(csv_filename) as f:
            reader = csv.reader(f)
            for row in reader:
                self.worksheet.append(row)

    def SetFoldersSheet(self, folders):
        # self.work_folder = excel_dir
        # self.wb = load_workbook(
        #     f'./{self.work_folder}/items.xlsm',
        #     keep_vba=True)
        # print(folders)
        # print(f'./{self.work_folder}/items.xlsm')
        self.worksheet = self.wb['folders']
        for row_idx, elem in enumerate(folders):
            print(elem)
            print(row_idx + 1)
            self.worksheet.cell(
                row=row_idx + 1, column=1).value = f"layer{row_idx+1}"
            self.worksheet.cell(row=row_idx + 1, column=2).value = elem

        self.wb.close()
